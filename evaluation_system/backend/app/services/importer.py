"""One deterministic CSV/XLSX importer shared by both evaluation modes."""

from __future__ import annotations

import csv
import hashlib
import io
import math
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any


class DatasetImportError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.public_message = message


@dataclass(frozen=True)
class ImportIssue:
    severity: str
    code: str
    message: str
    source_row_number: int | None = None
    field_name: str | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "code": self.code,
            "message": self.message,
            "source_row_number": self.source_row_number,
            "field_name": self.field_name,
        }


@dataclass
class ParsedTurn:
    source_row_number: int
    query: str
    source_session_id: str | None
    source_time_raw: str | None
    source_timestamp: datetime | None
    metadata: dict[str, Any] = field(default_factory=dict)
    turn_index: int = 0


@dataclass
class ParsedSession:
    source_session_id: str | None
    synthetic_session: bool
    synthetic_label: str | None
    turns: list[ParsedTurn]
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class ParsedDataset:
    filename: str | None
    source_type: str
    file_sha256: str | None
    dataset_type: str
    row_count: int
    valid_row_count: int
    invalid_row_count: int
    sessions: list[ParsedSession]
    issues: list[ImportIssue]
    metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def session_count(self) -> int:
        return len(self.sessions)

    def summary(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "file_sha256": self.file_sha256,
            "row_count": self.row_count,
            "valid_row_count": self.valid_row_count,
            "invalid_row_count": self.invalid_row_count,
            "session_count": self.session_count,
            "issues": [issue.as_dict() for issue in self.issues],
        }


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and value.strip().lower() in {"", "nan"}


def _text(value: Any) -> str | None:
    if _is_missing(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _timestamp(value: Any) -> tuple[str | None, datetime | None]:
    if _is_missing(value):
        return None, None
    if isinstance(value, datetime):
        parsed = value
        raw = value.isoformat()
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
        raw = value.isoformat()
    else:
        raw = str(value).strip()
        candidate = raw[:-1] + "+00:00" if raw.endswith("Z") else raw
        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError:
            try:
                parsed_time = time.fromisoformat(candidate)
            except ValueError:
                return raw, None
            # Imported timestamps control ordering only. Anchor a time-only
            # value to a fixed date so CSV and XLSX inputs sort identically
            # without introducing the current date into persisted data.
            parsed = datetime.combine(date(1970, 1, 1), parsed_time)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return raw, parsed.astimezone(timezone.utc)


def _csv_rows(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise DatasetImportError("INVALID_CSV_ENCODING", "CSV must be UTF-8 encoded") from exc
    sample = decoded[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
    return list(reader.fieldnames or []), list(reader)


def _xlsx_rows(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DatasetImportError(
            "XLSX_DEPENDENCY_MISSING", "XLSX support requires openpyxl"
        ) from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        headers = [str(value).strip() if value is not None else "" for value in (header_values or [])]
        rows = [dict(zip(headers, values)) for values in iterator]
        workbook.close()
    except Exception as exc:
        raise DatasetImportError("INVALID_XLSX", "XLSX workbook could not be parsed") from exc
    return headers, rows


def parse_dataset_file(
    *,
    filename: str,
    content: bytes,
    dataset_type: str,
    max_rows: int = 50_000,
) -> ParsedDataset:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        headers, rows = _csv_rows(content)
    elif suffix == ".xlsx":
        headers, rows = _xlsx_rows(content)
    else:
        raise DatasetImportError("UNSUPPORTED_FILE_TYPE", "Only CSV and XLSX are supported")
    return parse_dataset_rows(
        headers=headers,
        rows=rows,
        filename=Path(filename).name,
        file_sha256=hashlib.sha256(content).hexdigest(),
        dataset_type=dataset_type,
        max_rows=max_rows,
    )


def parse_dataset_rows(
    *,
    headers: list[str],
    rows: list[dict[str, Any]],
    filename: str | None,
    file_sha256: str | None,
    dataset_type: str,
    max_rows: int,
) -> ParsedDataset:
    normalized_headers = [str(header).strip().lower() for header in headers]
    if len(set(normalized_headers)) != len(normalized_headers):
        raise DatasetImportError("DUPLICATE_COLUMNS", "Column names must be unique")
    if "query" not in normalized_headers:
        raise DatasetImportError("MISSING_QUERY_COLUMN", "Required column 'query' is missing")
    if len(rows) > max_rows:
        raise DatasetImportError("ROW_LIMIT_EXCEEDED", f"Dataset exceeds {max_rows} rows")

    key_map = dict(zip(normalized_headers, headers))
    issues: list[ImportIssue] = []
    valid: list[ParsedTurn] = []
    for source_row, row in enumerate(rows, start=2):
        query = _text(row.get(key_map["query"]))
        if not query:
            issues.append(ImportIssue(
                severity="ERROR", code="EMPTY_QUERY", message="Query is empty",
                source_row_number=source_row, field_name="query",
            ))
            continue
        session_id = _text(row.get(key_map.get("session_id"))) if "session_id" in key_map else None
        raw_time, parsed_time = _timestamp(row.get(key_map.get("time"))) if "time" in key_map else (None, None)
        if raw_time is not None and parsed_time is None:
            issues.append(ImportIssue(
                severity="WARNING", code="INVALID_TIMESTAMP",
                message="Timestamp is invalid; this session preserves source row order",
                source_row_number=source_row, field_name="time",
            ))
        valid.append(ParsedTurn(
            source_row_number=source_row,
            query=query,
            source_session_id=session_id,
            source_time_raw=raw_time,
            source_timestamp=parsed_time,
        ))

    grouped: dict[str, list[ParsedTurn]] = {}
    session_order: list[str] = []
    session_properties: dict[str, tuple[str | None, bool, str | None]] = {}
    for turn in valid:
        if turn.source_session_id is None:
            key = f"__synthetic_row_{turn.source_row_number}"
            props = (None, True, f"synthetic-row-{turn.source_row_number}")
        else:
            key = f"source:{turn.source_session_id}"
            props = (turn.source_session_id, False, None)
        if key not in grouped:
            grouped[key] = []
            session_order.append(key)
            session_properties[key] = props
        grouped[key].append(turn)

    sessions: list[ParsedSession] = []
    for key in session_order:
        turns = grouped[key]
        has_all_timestamps = all(turn.source_timestamp is not None for turn in turns)
        if has_all_timestamps:
            turns.sort(key=lambda turn: (turn.source_timestamp, turn.source_row_number))
        else:
            turns.sort(key=lambda turn: turn.source_row_number)
            code = "NO_VALID_SESSION_TIMESTAMPS" if all(
                turn.source_timestamp is None for turn in turns
            ) else "INCOMPLETE_SESSION_TIMESTAMPS"
            issues.append(ImportIssue(
                severity="WARNING", code=code,
                message="Session preserves source row ordering because timestamps are incomplete",
                source_row_number=turns[0].source_row_number,
                field_name="time",
            ))
        for index, turn in enumerate(turns, start=1):
            turn.turn_index = index
        source_id, synthetic, label = session_properties[key]
        sessions.append(ParsedSession(
            source_session_id=source_id,
            synthetic_session=synthetic,
            synthetic_label=label,
            turns=turns,
            metadata={"ordering": "timestamp_then_row" if has_all_timestamps else "source_row"},
        ))

    return ParsedDataset(
        filename=filename,
        source_type="FILE",
        file_sha256=file_sha256,
        dataset_type=dataset_type,
        row_count=len(rows),
        valid_row_count=len(valid),
        invalid_row_count=len(rows) - len(valid),
        sessions=sessions,
        issues=issues,
        metadata={"parser_version": "evaluation-importer-v1", "columns": normalized_headers},
    )


def parse_manual_dataset(
    queries: list[str], *, dataset_type: str = "STABILITY"
) -> ParsedDataset:
    issues: list[ImportIssue] = []
    turns: list[ParsedTurn] = []
    for index, value in enumerate(queries, start=1):
        query = str(value).strip()
        if not query:
            issues.append(ImportIssue(
                severity="ERROR", code="EMPTY_QUERY", message="Query is empty",
                source_row_number=index, field_name="query",
            ))
            continue
        turns.append(ParsedTurn(
            source_row_number=index, query=query, source_session_id=None,
            source_time_raw=None, source_timestamp=None, turn_index=len(turns) + 1,
        ))
    if not turns:
        raise DatasetImportError("NO_VALID_QUERIES", "At least one non-empty query is required")
    return ParsedDataset(
        filename=None, source_type="MANUAL", file_sha256=None,
        dataset_type=dataset_type, row_count=len(queries),
        valid_row_count=len(turns), invalid_row_count=len(queries) - len(turns),
        sessions=[ParsedSession(
            source_session_id=None, synthetic_session=True,
            synthetic_label="manual-session", turns=turns,
            metadata={"ordering": "manual"},
        )],
        issues=issues,
        metadata={"parser_version": "evaluation-importer-v1", "manual": True},
    )
